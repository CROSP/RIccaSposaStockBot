from __future__ import annotations
import io, json, hashlib, pathlib, re, requests, logging
import uuid
from typing import Any, Dict, List, Tuple, Callable
from collections import defaultdict
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from PIL import Image
from openpyxl.worksheet.hyperlink import Hyperlink
from io import BytesIO

# ────────────────────────────────────────────────────────────
# LOGGING SETUP
# ────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[logging.FileHandler("stock_builder.log"), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# ────────────────────────────────────────────────────────────
# CONSTANTS & REGEX
# ────────────────────────────────────────────────────────────
CACHE_DIR = pathlib.Path(".img_cache");
CACHE_DIR.mkdir(exist_ok=True)

SIZE_RE = re.compile(r"\b(\d{2}(?:\.0)?|XS|S|M|L|XL|XXL|XXXL)\b", re.I)

# 🔸  SKU_RE now also matches a trailing "‑number" that we'll normalise later
SKU_RE = re.compile(
    r"([A-Za-z]{0,2}\d{1,2}[-–][A-Za-z]?[-–]?\d{3}(?:[-–]\d+|\(\d+\))?)"
)

COMPOSITE_RE = re.compile(
    r"(?P<sku>(?:[A-Za-z]{0,2}\d{1,2}-\d{3})(?:-\d+|\(\d+\))?)"
    r".*?\(\s*(?P<size>[^,;)\s]+)"
    r"(?:\s*,\s*(?P<color>[^;)\s]+))?"
    r"(?:;\s*(?P<note>[^)]+))?\)",
    re.I | re.S,
)

COLOR_MAP = {
    "айворі": "Ivory", "ivory": "Ivory",
    "оригінал": "Original", "оригинал": "Original",
    "white": "White", "milk": "Milk", "cream": "Cream",
    "champagne": "Champagne", "nude": "Nude", "blush": "Blush",
}


# ────────────────────────────────────────────────────────────
# SMALL HELPERS
# ────────────────────────────────────────────────────────────
def first(it, default=""):
    return next((x for x in it if x), default)


def to_float(v: str | None) -> float | None:
    if not v:
        return None
    s = (
        str(v)
        .replace("\u202f", "")
        .replace("\xa0", "")
        .replace(" ", "")
        .replace(",", ".")
    )
    try:
        return float(s)
    except ValueError:
        return None


# 🔸  NEW robust normaliser -------------------------------------------------
_VARIANT_RE = re.compile(r"^([A-Za-z]{0,2}\d{1,2}-\d{3})-(\d+)$")  # 26-012-3
_PAREN_RE = re.compile(r"^([A-Za-z]{0,2}\d{1,2}-\d{3})\((\d+)\)$")  # 26-012(3)


def normalize_sku(txt: str) -> str:
    """
    • Unify dashes, trim, lowercase.
    • Convert dashed variant "26‑012‑3" → parenthesised "26‑012(3)".
    • Leave an already‑parenthesised SKU untouched.
    """
    if not txt:
        return ""
    t = txt.strip().replace("_", "-").replace("–", "-")
    m_dash = _VARIANT_RE.match(t)
    if m_dash:
        t = f"{m_dash.group(1)}({m_dash.group(2)})"
    # keep an existing parenthesised form as‑is
    t = re.sub(r"[^A-Za-z0-9\-\(\)]", "", t)  # allow () now
    t = re.sub(r"-{2,}", "-", t)
    return t.lower()


# --------------------------------------------------------------------------


def accessories_string(acc: list[dict[str, Any]], cur: str) -> str:
    return ", ".join(
        f"{a['accessories_name']} {a['prices'][cur]}"
        for a in acc
        if cur in a["prices"]
    ) if acc else ""


def accessories_sum(acc: list[dict[str, Any]], cur: str) -> float:
    return sum(
        to_float(a["prices"].get(cur)) or 0
        for a in acc
        if cur in a["prices"]
    )


def _png_bytes_for_xlsx(url: str) -> bytes:
    """Download an image from URL and convert it to PNG bytes"""
    r = requests.get(url, timeout=20)
    img = Image.open(BytesIO(r.content))
    if img.mode in ("P", "RGBA"):
        img = img.convert("RGB")
    out = BytesIO()
    img.save(out, format="PNG")
    return out.getvalue()  # Return actual bytes


def cached_image_bytes(url: str) -> bytes:
    """Return image bytes with caching to avoid redundant downloads"""
    p = CACHE_DIR / (hashlib.md5(url.encode()).hexdigest() + ".img")
    if p.exists():
        return p.read_bytes()  # Return actual bytes from cache

    # Get fresh data as bytes
    data = _png_bytes_for_xlsx(url)
    p.write_bytes(data)  # Write bytes to cache
    return data  # Return the actual bytes

def sku_sort_key(s: str):
    return [int(x) for x in re.split(r"[^\d]", s) if x.isdigit()]


def parse_composite_blob(blob: str):
    m = COMPOSITE_RE.search(blob)
    if not m:
        return "", "", "", ""
    return (
        m.group("sku"),
        m.group("size"),
        m.group("color") or "",
        m.group("note") or "",
    )


# Function to determine discount based on year in SKU
def get_discount_by_sku_year(sku: str) -> float:
    """Determine discount percentage based on the year in the SKU"""
    if not sku:
        return 20.0  # Default to 2025 discount if no SKU

    # Extract the year from SKU (first two digits represent the year)
    year_match = re.match(r'^[A-Za-z]{0,2}(\d{1,2})-', sku)
    if year_match:
        year_digits = year_match.group(1)
        if year_digits.startswith('24'):
            return 30.0  # 30% discount for 2024

    # Default to 20% for 2025 and any other year
    return 20.0


# ────────────────────────────────────────────────────────────
# SHEET-SPECIFIC ROW MAPPERS  (sku, size, color, note)
# ────────────────────────────────────────────────────────────
def map_row_simple(idx=(1, 2, 3, 4)):
    s_i, sz_i, c_i, n_i = idx

    def f(v: list[str]): return (
        v[s_i] if len(v) > s_i else "", v[sz_i] if len(v) > sz_i else "",
        v[c_i] if len(v) > c_i else "", v[n_i] if len(v) > n_i else "")

    return f


def map_row_composite(col=1, note_ok=True):
    """
    Extract (sku, size, color, note) from a composite string format in the specified column.
    Handles both standard formats and variations with different delimiters.
    """

    def f(v: list[str]):
        text = v[col] if len(v) > col else ""

        # Try to match with the COMPOSITE_RE pattern
        m = COMPOSITE_RE.search(text)
        if m:
            return (
                m.group("sku"),
                m.group("size"),
                m.group("color") or "",  # Add or "" to handle None
                (m.group("note") or "") if note_ok else ""
            )

        # If no match, try a simpler approach to extract just the SKU
        sku_match = SKU_RE.search(text)
        if sku_match:
            sku = sku_match.group(1)
            # Try to extract size and color from parentheses if present
            paren_match = re.search(r"\(\s*([^,;)\s]+)(?:\s*,\s*([^;)\s]+))?(?:;\s*([^)]+))?\)", text)
            if paren_match:
                size = paren_match.group(1) or ""
                color = paren_match.group(2) or ""  # Add or "" to handle None
                note = paren_match.group(3) or "" if note_ok else ""
                return sku, size, color, note

            return sku, "", "", ""

        return "", "", "", ""

    return f


def map_row_kansas(v: list[str]):
    sku = v[2] if len(v) > 2 else ""
    parts = [p.strip() for p in (v[4] if len(v) > 4 else "").split(",")]
    return sku, (parts[0] if parts else ""), (parts[1] if len(parts) > 1 else ""), ""


def map_row_soy_unica(v: list[str]):
    return (v[1] if len(v) > 1 else "", "", v[2] if len(v) > 2 else "", "")


ROW_MAPPERS: Dict[str, Callable[[list[str]], Tuple[str, str, str, str]]] = {
    "Romania BucharestWarehouse": map_row_simple(),
    "Romania BucharestStore": map_row_composite(1),
    "Ukraine Warehouse": map_row_composite(1, note_ok=False),
    "Ukraine WarehouseAfterBarcelona": map_row_simple((1, 2, 3, 4)),
    "England HAMMINGBRIDAL": map_row_simple(),
    "England Katya": map_row_simple((1, 2, 3, 4)),
    "USA Vera": map_row_simple(), "USA Pica": map_row_simple(),
    "Canada ROYAL": map_row_composite(1, note_ok=False),
    "USA Kansas": map_row_kansas, "USA Kimberley": map_row_simple(),
    "SOY UNICA": map_row_soy_unica,
    # New mappings
    "USA Melissa PR": map_row_simple(),
    "USA Pickyglam": map_row_simple(),
    "Mexico SOY UNICA": map_row_soy_unica,
}


def _build_catalogue_index(cat: list[dict]) -> Dict[str, dict]:
    idx: Dict[str, dict] = {}
    for item in cat:
        norm = normalize_sku(item["sku"])
        idx[norm] = item
        # 🔸 add alt key if the raw SKU is dashed & convertible
        m = _VARIANT_RE.match(item["sku"])
        if m:
            idx[normalize_sku(f"{m.group(1)}({m.group(2)})")] = item
    return idx


# ────────────────────────────────────────────────────────────
# CORE BUILDER
# ────────────────────────────────────────────────────────────
def build_stock_excel(stock_xlsx: str, catalogue_json: str,
                      output_xlsx: str, thumb_px: int = 140):
    logger.info("Starting stock building process")

    # Load catalogue data
    cat = json.load(open(catalogue_json, encoding="utf-8"))
    cat_by_sku = {normalize_sku(i["sku"]): i for i in cat}
    logger.info(f"Loaded catalogue with {len(cat)} items")

    # Get total stock items before processing
    xls = pd.ExcelFile(stock_xlsx)
    raw_count = 0
    stock_by_sheet = {}

    # First pass to count all raw items
    for sheet in xls.sheet_names:
        df = xls.parse(sheet, header=None).fillna("")
        sheet_count = len(df.iloc[1:])  # Excluding header row
        stock_by_sheet[sheet] = sheet_count
        raw_count += sheet_count

    logger.info(f"TOTAL RAW STOCK COUNT: {raw_count} items across {len(xls.sheet_names)} sheets")
    logger.info("Raw counts by sheet:")
    for sheet, count in stock_by_sheet.items():
        logger.info(f"  {sheet}: {count} items")

    # Statistics containers
    total_stock_items = 0
    not_found_items = []
    invalid_sku_items = []  # NEW: Track items with invalid SKUs
    location_stats = defaultdict(int)
    stock_by_country = defaultdict(int)

    rows: List[Dict[str, Any]] = []

    # Process each sheet
    logger.info("Processing stock file in detail")
    valid_sku_count = 0
    skipped_items = 0

    for sheet in xls.sheet_names:
        country, *loc = sheet.split(maxsplit=1)
        location = loc[0] if loc else ""
        df = xls.parse(sheet, header=None).fillna("")
        mapper = ROW_MAPPERS.get(sheet.strip())

        sheet_items = 0
        sheet_not_found = 0
        sheet_valid_skus = 0
        sheet_skipped = 0

        for idx, r in df.iloc[1:].iterrows():
            vals = [str(v).strip() for v in r.tolist()]
            blob = " ".join(vals)
            total_stock_items += 1
            sheet_items += 1

            if mapper:
                sku_raw, size_txt, color_txt, note_txt = mapper(vals)
            else:
                sku_raw = size_txt = color_txt = note_txt = ""

            # fall back / fix bogus size-as-sku
            if not SKU_RE.fullmatch(sku_raw or ""):
                sku_raw, size_txt, color_txt, note_txt = parse_composite_blob(blob)

            # Skip rows without valid SKUs - but track them for reporting
            if not SKU_RE.fullmatch(sku_raw or ""):
                skipped_items += 1
                sheet_skipped += 1

                # NEW: Save the invalid SKU item for reporting
                invalid_sku_items.append({
                    "sheet": sheet,
                    "country": country,
                    "location": location,
                    "row_index": idx + 2,  # Excel row number (add 2 for header and 0-indexing)
                    "raw_data": blob[:100] + ("..." if len(blob) > 100 else ""),  # First 100 chars of raw data
                    "attempted_sku": sku_raw,
                })
                continue

            valid_sku_count += 1
            sheet_valid_skus += 1

            sku_norm = normalize_sku(sku_raw)
            size = size_txt.upper()
            color = COLOR_MAP.get(color_txt.lower() if color_txt else "", color_txt or "")

            cat_item = cat_by_sku.get(sku_norm, {})

            # Log not found items
            if not cat_item:
                not_found_items.append({
                    "sku": sku_raw,
                    "normalized_sku": sku_norm,
                    "country": country,
                    "location": location,
                    "size": size,
                    "color": color,
                })
                sheet_not_found += 1
                continue

            pd_r = cat_item.get("price_data", {})
            if isinstance(pd_r, list): pd_r = next((d for d in pd_r if isinstance(d, dict)), {})
            if not isinstance(pd_r, dict): pd_r = {}

            price = pd_r.get("price", {});
            acc = pd_r.get("accessories", [])

            # Get discount based on SKU year
            discount_val = get_discount_by_sku_year(sku_raw)

            usd = to_float(price.get("usd"));
            eur = to_float(price.get("eur"));
            gbp = to_float(price.get("gbp"))
            acc_usd = accessories_sum(acc, "usd");
            acc_eur = accessories_sum(acc, "eur");
            acc_gbp = accessories_sum(acc, "gbp")

            # Calculate total prices (without discount) and total with discount
            total_usd = (usd or 0) + acc_usd if usd is not None else None
            total_eur = (eur or 0) + acc_eur if eur is not None else None
            total_gbp = (gbp or 0) + acc_gbp if gbp is not None else None

            total_usd_with_discount = total_usd - discount_val if total_usd is not None else None
            total_eur_with_discount = total_eur - discount_val if total_eur is not None else None
            total_gbp_with_discount = total_gbp - discount_val if total_gbp is not None else None

            stock_by_country[country] += 1
            location_stats[f"{country} - {location}"] += 1

            rows.append({
                "NameSKU": f"{(pd_r.get('modelName') or cat_item.get('title', '')).strip()} {cat_item.get('sku', sku_raw)}",
                "SKU": cat_item.get("sku", sku_raw),
                "Collection": cat_item.get("collection", ""),
                "Size": size, "Color": color or pd_r.get("color", ""),
                "Qty": 1,  # NEW: Added Qty column with default value 1
                "Country": country, "Location": location, "Notes": note_txt,
                "Photo": first(cat_item.get("images", []), ""),
                "Link": cat_item.get("detailed_photos_link", ""),
                "Dress_USD": price.get("usd", ""), "Accessories_USD": accessories_string(acc, "usd"),
                "Total_USD": total_usd if total_usd is not None else "",
                "Total_USD_Discounted": total_usd_with_discount if total_usd_with_discount is not None else "",
                "Dress_EUR": price.get("eur", ""), "Accessories_EUR": accessories_string(acc, "eur"),
                "Total_EUR": total_eur if total_eur is not None else "",
                "Total_EUR_Discounted": total_eur_with_discount if total_eur_with_discount is not None else "",
                "Dress_GBP": price.get("gbp", ""), "Accessories_GBP": accessories_string(acc, "gbp"),
                "Total_GBP": total_gbp if total_gbp is not None else "",
                "Total_GBP_Discounted": total_gbp_with_discount if total_gbp_with_discount is not None else "",
                "Discount": str(discount_val) + "%" , "SKU_norm": sku_norm,
                "Year": "2024" if discount_val == 30.0 else "2025",  # NEW: Added Year column for clarity
            })

        logger.info(
            f"Sheet '{sheet}': {sheet_items} total items, {sheet_valid_skus} valid SKUs, {sheet_not_found} not found in catalogue, {sheet_skipped} skipped (invalid SKU)")

    # Log summary statistics
    total_found = valid_sku_count - len(not_found_items)
    logger.info("=" * 60)
    logger.info("PROCESSING SUMMARY:")
    logger.info(f"Total raw stock items: {raw_count}")
    logger.info(f"Total items processed: {total_stock_items}")
    logger.info(f"Items with valid SKUs: {valid_sku_count} ({valid_sku_count / total_stock_items * 100:.1f}%)")
    logger.info(f"Items skipped (invalid SKUs): {skipped_items} ({skipped_items / total_stock_items * 100:.1f}%)")
    logger.info(f"Valid items found in catalogue: {total_found} ({total_found / valid_sku_count * 100:.1f}%)")
    logger.info(
        f"Valid items not found in catalogue: {len(not_found_items)} ({len(not_found_items) / valid_sku_count * 100:.1f}%)")
    logger.info("=" * 60)

    # Log discount statistics
    discount_stats = {
        "30.0%": len([r for r in rows if r["Discount"] == 30.0]),
        "20.0%": len([r for r in rows if r["Discount"] == 20.0])
    }
    logger.info("Discount statistics:")
    logger.info(f"  2024 items (30% discount): {discount_stats['30.0%']} items")
    logger.info(f"  2025 items (20% discount): {discount_stats['20.0%']} items")

    # NEW: Log invalid SKU items
    if invalid_sku_items:
        logger.warning(f"Found {len(invalid_sku_items)} items with invalid SKUs:")
        for i, item in enumerate(invalid_sku_items[:20], 1):  # Show first 20 items
            logger.warning(
                f"  {i}. Sheet: {item['sheet']}, Row: {item['row_index']}, Attempted SKU: '{item['attempted_sku']}', Raw Data: {item['raw_data']}"
            )

        if len(invalid_sku_items) > 20:
            logger.warning(f"  ... and {len(invalid_sku_items) - 20} more. See invalid_skus.xlsx for complete list.")

    # Log country statistics
    logger.info("Stock by country:")
    for country, count in sorted(stock_by_country.items(), key=lambda x: x[1], reverse=True):
        logger.info(f"  {country}: {count} items")

    # Log location statistics
    logger.info("Stock by location:")
    for location, count in sorted(location_stats.items(), key=lambda x: x[1], reverse=True):
        logger.info(f"  {location}: {count} items")

    # Log not found dresses
    if not_found_items:
        logger.warning("The following SKUs were not found in the catalogue:")
        for item in not_found_items:
            logger.warning(
                f"  SKU: {item['sku']} (norm: {item['normalized_sku']}) - {item['country']}, {item['location']}, Size: {item['size']}, Color: {item['color']}")

    # Sort by Year descending (2025 first, then 2024)
    rows.sort(key=lambda r: (r["Year"], r["Collection"].lower(), sku_sort_key(r["SKU_norm"])), reverse=True)

    cols = ["NameSKU", "SKU", "Collection", "Size", "Color", "Qty", "Country", "Location", "Notes",
            "Photo", "Link",
            "Dress_USD", "Accessories_USD", "Total_USD", "Total_USD_Discounted",
            "Dress_EUR", "Accessories_EUR", "Total_EUR", "Total_EUR_Discounted",
            "Dress_GBP", "Accessories_GBP", "Total_GBP", "Total_GBP_Discounted",
            "Discount", "Year"]

    df = pd.DataFrame(rows)
    df.drop(columns=["SKU_norm"], errors="ignore", inplace=True)
    df = df[cols]
    df.to_excel(output_xlsx, index=False)
    logger.info(f"Created Excel file with {len(df)} processed items")

    # ─── POST-PROCESS EXCEL ─────────────────────────────────
    wb = load_workbook(output_xlsx)
    ws = wb.active
    wrap = Alignment(wrap_text=True, vertical="top")
    for col in ws.iter_cols():
        for cell in col: cell.alignment = wrap

    # widen Location & Photo columns
    ws.column_dimensions[get_column_letter(cols.index("Location") + 1)].width = 22
    photo_col = cols.index("Photo") + 1
    ws.column_dimensions[get_column_letter(photo_col)].width = round(thumb_px / 7, 1)

    # NEW: Add helper columns for URLs
    # We'll add these at the end of the sheet to keep URLs accessible to the bot
    helper_cols_start = len(cols) + 1
    ws.cell(row=1, column=helper_cols_start, value="Photo_URL")
    ws.cell(row=1, column=helper_cols_start + 1, value="Link_URL")

    # widen the Link column
    link_col = cols.index("Link") + 1
    ws.column_dimensions[get_column_letter(link_col)].width = 30

    # hyperlink text
    for i, url in enumerate(df["Link"], start=2):
        if url:
            # Create the visible hyperlink cell
            cell = ws.cell(row=i, column=link_col)
            cell.value = "ALL PHOTOS LINK"
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")

            # Store raw URL in helper column for TG bot to find
            ws.cell(row=i, column=helper_cols_start + 1, value=url)

    # thumbnails & row height
    row_h_pt = 170 * 0.75
    image_errors = 0
    for i, url in enumerate(df["Photo"], start=2):
        ws.row_dimensions[i].height = row_h_pt
        cell = ws.cell(row=i, column=photo_col)

        # Clear the cell to prepare for image
        cell.value = None

        # Store URL in helper column for TG bot
        ws.cell(row=i, column=helper_cols_start, value=url)

        if not url: continue
        try:
            # Get the image bytes
            img_bytes = cached_image_bytes(url)

            # Create a new BytesIO object for PIL to use
            img_io = BytesIO(img_bytes)
            img = Image.open(img_io)

            # Calculate cell dimensions in pixels (approximate)
            cell_width_px = thumb_px
            cell_height_px = row_h_pt * 1.33  # Convert points to pixels (approx)

            # Resize image to fill cell while maintaining aspect ratio
            img_width, img_height = img.size
            width_ratio = cell_width_px / img_width
            height_ratio = cell_height_px / img_height

            # Use the larger ratio to ensure the image fills the cell
            resize_ratio = max(width_ratio, height_ratio)
            new_width = int(img_width * resize_ratio)
            new_height = int(img_height * resize_ratio)

            # Crop to cell dimensions if needed
            img = img.resize((new_width, new_height), Image.LANCZOS)

            # If image is larger than cell, crop from center
            if new_width > cell_width_px or new_height > cell_height_px:
                left = (new_width - cell_width_px) // 2 if new_width > cell_width_px else 0
                top = (new_height - cell_height_px) // 2 if new_height > cell_height_px else 0
                right = left + cell_width_px if new_width > cell_width_px else new_width
                bottom = top + cell_height_px if new_height > cell_height_px else new_height
                img = img.crop((left, top, right, bottom))

            # Create a fresh BytesIO object for the processed image
            img_io_final = BytesIO()
            img.save(img_io_final, format="PNG")
            img_io_final.seek(0)

            # Create the Excel image object
            xl_img = XLImage(img_io_final)
            xl_img.anchor = cell.coordinate
            xl_img._id = uuid.uuid4().hex  # unique inside ZIP
            ws.add_image(xl_img)
        except Exception as e:
            image_errors += 1
            logger.error(f"Thumbnail error for {url}: {e}")
            # Add more detailed logging
            import traceback
            logger.error(f"Detailed error: {traceback.format_exc()}")

    if image_errors:
        logger.warning(f"Encountered {image_errors} thumbnail errors")

    wb.save(output_xlsx)
    logger.info(f"✅ Stock file written → {output_xlsx}")

    # Generate missing SKUs report
    if not_found_items:
        missing_df = pd.DataFrame(not_found_items)
        missing_output = output_xlsx.replace(".xlsx", "_missing_skus.xlsx")
        missing_df.to_excel(missing_output, index=False)
        logger.info(f"✅ Missing SKUs file written → {missing_output}")

    # NEW: Generate invalid SKUs report
    if invalid_sku_items:
        invalid_skus_df = pd.DataFrame(invalid_sku_items)
        invalid_skus_output = output_xlsx.replace(".xlsx", "_invalid_skus.xlsx")
        invalid_skus_df.to_excel(invalid_skus_output, index=False)
        logger.info(f"✅ Invalid SKUs file written → {invalid_skus_output}")


# ────────────────────────────────────────────────────────────
# CLI
# ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse, sys

    ap = argparse.ArgumentParser(description="Generate Ricca in-stock workbook")
    ap.add_argument("stock", help="Path to raw stock .xlsx")
    ap.add_argument("catalogue", help="Path to JSON catalogue")
    ap.add_argument("-o", "--output", default="stock_result.xlsx")
    ap.add_argument("--log-level", choices=["DEBUG", "INFO", "WARNING", "ERROR"], default="INFO",
                    help="Set the logging level")
    ns = ap.parse_args()
    try:
        # Set log level from command line
        logger.setLevel(getattr(logging, ns.log_level))
        logger.info(f"Log level set to {ns.log_level}")

        build_stock_excel(ns.stock, ns.catalogue, ns.output)
    except Exception as err:
        logger.exception("Error in stock builder process")
        sys.exit(f"❌  {err}")