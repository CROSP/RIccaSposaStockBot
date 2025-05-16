"""
Telegram bot for processing stock_result.xlsx files.
Allows users to select columns, filter by countries, collections, and years,
and export as XLSX or PDF. Can upload files or use existing files.

‼️ 2025‑05 update
    • Keeps `Photo` and `Link` columns internally so thumbnails & hyperlinks are
      never lost, even if the user hides them.
    • Optionally hides those columns in the final workbook if the user removed
      them from the visible selection.
    • Robust hyperlink writing with `write_url()` (no silent drops).
    • Improved thumbnail handling with better image caching and compression.
    • File size management to stay under Telegram's limits.
"""
from __future__ import annotations

import hashlib
# ── stdlib ────────────────────────────────────────────────────────────────
from copy import deepcopy
import io
import json
import logging
import os
import re
import shutil
import subprocess
import tempfile
from datetime import datetime
from distutils.version import LooseVersion
from functools import partial
from glob import glob
from io import BytesIO
from pathlib import Path
from typing import Any, List, Optional
from urllib.parse import urlparse

# ── third‑party ────────────────────────────────────────────────────────────
import openpyxl  # only to open existing workbooks (never for writing here)
import pandas as pd
import requests
import xlsxwriter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image
from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    Update,
)
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ConversationHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

# ── logging ───────────────────────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ── Conversation states ───────────────────────────────────────────────────
(
    SELECT_SOURCE,
    SELECT_FILE,
    SELECT_EXISTING_FILE,
    SELECTING_COLUMNS,
    CONFIRM_COLUMNS,
    FILTER_YEARS,
    CONFIRM_YEARS,
    FILTER_COUNTRIES,
    CONFIRM_COUNTRIES,
    FILTER_COLLECTIONS,
    CONFIRM_COLLECTIONS,
    SELECT_FORMAT,
) = range(12)

# Default column selection (all checked except accessories, price per dress, location)
DEFAULT_EXCLUDED_COLUMNS = [
    "Accessories_USD",
    "Accessories_EUR",
    "Accessories_GBP",
    "Dress_USD",
    "Dress_EUR",
    "Dress_GBP",
    "Location",
]

DATA_DIR = Path(__file__).resolve().parent
MAX_FILE_SIZE = 45 * 1024 * 1024  # 45 MB

# ────────────────────────────────────────────────────────────────
# Helper functions
# ────────────────────────────────────────────────────────────────

def _find_existing_files() -> List[str]:
    return sorted(
        str(p) for p in DATA_DIR.glob("*.*") if p.suffix.lower() in (".xlsx", ".json")
    )


def _recover_photo_column(df: pd.DataFrame) -> pd.DataFrame:
    if "Photo" not in df.columns:
        return df
    photo_url_col = next((c for c in df.columns if str(c).lower() == "photo_url"), None)
    if photo_url_col is not None:
        df["Photo"] = df[photo_url_col]
        return df
    photo_idx = df.columns.get_loc("Photo")
    if photo_idx + 1 < len(df.columns):
        helper = df.columns[photo_idx + 1]
        if not helper or str(helper).startswith("Unnamed"):
            df["Photo"] = df[helper]
            df.drop(columns=[helper], inplace=True)
    return df


def _recover_link_column(df: pd.DataFrame, file_path: str) -> pd.DataFrame:
    if "Link" not in df.columns:
        return df
    link_url_col = next((c for c in df.columns if str(c).lower() == "link_url"), None)
    if link_url_col is not None:
        df["Link"] = df[link_url_col]
        return df
    link_idx = df.columns.get_loc("Link")
    if link_idx + 1 < len(df.columns):
        helper = df.columns[link_idx + 1]
        if not helper or str(helper).startswith("Unnamed"):
            df["Link"] = df[helper]
            return df
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=link_idx + 1)
            if cell.hyperlink:
                df.iat[r - 2, link_idx] = cell.hyperlink.target
        wb.close()
    except Exception:
        pass
    return df


CACHE_DIR = Path(__file__).with_suffix(".img_cache"); CACHE_DIR.mkdir(exist_ok=True)


def cached_png(url: str, *, compression_level: int = 6) -> Optional[BytesIO]:
    if not url or not url.startswith("http"):
        return None
    # Use compression level in cache key to store different versions
    cache_key = f"{hashlib.md5(url.encode()).hexdigest()}_{compression_level}.png"
    p = CACHE_DIR / cache_key
    if p.exists():
        try:
            bio = BytesIO(p.read_bytes()); bio.seek(0); return bio
        except Exception:
            pass
    try:
        r = requests.get(url, timeout=10); r.raise_for_status()
        img = Image.open(BytesIO(r.content))
        if img.mode in ("P", "RGBA"):
            img = img.convert("RGB")
        w, h = img.size
        # Make smaller images for higher compression
        max_height = 160 if compression_level >= 8 else 200
        if h > max_height:
            w = int(w * max_height / h); img = img.resize((w, max_height), Image.LANCZOS)
        bio = BytesIO(); img.save(bio, "PNG", optimize=True, compress_level=compression_level)
        bio.seek(0); p.write_bytes(bio.getbuffer()); bio.seek(0); return bio
    except Exception as exc:
        logger.error(f"img fetch fail {url}: {exc}"); return None

def insert_image(ws, row: int, col: int, bio: Optional[BytesIO]) -> None:
    if bio is None:
        return
    try:
        if LooseVersion(xlsxwriter.__version__) >= LooseVersion("1.2.0"):
            bio.seek(0)
            ws.insert_image(row, col, "thumb.png", {"image_data": bio, "x_scale": 0.8, "y_scale": 0.8})
        else:
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                bio.seek(0); shutil.copyfileobj(bio, tmp); temp_path = tmp.name
            ws.insert_image(row, col, temp_path, {"x_scale": 0.8, "y_scale": 0.8})
            os.unlink(temp_path)
    except Exception as exc:
        logger.error(f"insert img fail: {exc}")


def split_dataframe(df: pd.DataFrame, *, max_rows: int = 1000) -> List[pd.DataFrame]:
    return [df.iloc[i : i + max_rows] for i in range(0, len(df), max_rows)]

# ────────────────────────────────────────────────────────────────
# Helper
# ────────────────────────────────────────────────────────────────
def insert_image_compat(ws, row, col, img_bio, *, x_scale=0.55, y_scale=0.55):
    """
    Insert a PNG from BytesIO into a worksheet, whether or not
    the XlsxWriter build supports the 'image_data' keyword.

    Returns True if successful, False otherwise.
    """
    if img_bio is None:
        return False

    try:
        # First try with image_data parameter
        if LooseVersion(xlsxwriter.__version__) >= LooseVersion("1.2.0"):
            # Make sure to reset the position to the beginning of the BytesIO object
            img_bio.seek(0)
            ws.insert_image(row, col, "thumb.png",
                            {"image_data": img_bio, "x_scale": x_scale, "y_scale": y_scale})
        else:
            # For older versions, write to a temporary file
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                # Copy all data from the BytesIO to the temp file
                img_bio.seek(0)
                tmp.write(img_bio.read())
                tmp_path = tmp.name
            ws.insert_image(row, col, tmp_path,
                            {"x_scale": x_scale, "y_scale": y_scale})
            os.unlink(tmp_path)
        return True
    except Exception as e:
        logger.error(f"Failed to insert image: {str(e)}")
        return False

def get_file_size(file_path: str) -> int:
    """Get the size of a file in bytes"""
    return os.path.getsize(file_path)


# ────────────────────────────────────────────────────────────────
# Bot Class
# ────────────────────────────────────────────────────────────────
class StockSelectorBot:
    """Bot that handles stock file selection and filtering."""

    def __init__(self):
        self.df: pd.DataFrame | None = None
        self.file_path: str | None = None
        # selections
        self.selected_columns: list[str] = []
        self.all_columns: list[str] = []
        self.selected_years: list[int] = []
        self.all_years: list[int] = []
        self.selected_countries: list[str] = []
        self.all_countries: list[str] = []
        self.selected_collections: list[str] = []
        self.all_collections: list[str] = []
        self.output_format: str = "xlsx"
        self.existing_files: list[str] = []

    # ── entry ────────────────────────────────────────────────────
    async def start(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        await update.message.reply_text(
            "Welcome to the Stock Selector Bot! 📊\n\n"
            "I'll help you filter and format your stock data."
        )
        return await self.show_source_selection(update, context)

    # ── source pick ──────────────────────────────────────────────
    async def show_source_selection(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        self.existing_files = _find_existing_files()
        kb: list[list[InlineKeyboardButton]] = [
            [InlineKeyboardButton("📤 Upload New File", callback_data="source_upload")]
        ]
        if self.existing_files:
            kb.append([InlineKeyboardButton("📁 Use Existing Files", callback_data="source_existing")])
        markup = InlineKeyboardMarkup(kb)
        if update.message:
            await update.message.reply_text("How would you like to provide your stock data?", reply_markup=markup)
        else:
            await update.callback_query.edit_message_text("How would you like to provide your stock data?", reply_markup=markup)
        return SELECT_SOURCE

    async def select_source(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        query = update.callback_query; await query.answer()
        src = query.data.removeprefix("source_")
        if src == "upload":
            await query.edit_message_text("Please upload your *stock_result.xlsx* file to begin.")
            return SELECT_FILE
        if src == "existing":
            return await self.show_existing_files(update, context)
        await query.edit_message_text("Invalid selection. Please try again.")
        return await self.show_source_selection(update, context)

    # ── existing files ───────────────────────────────────────────
    async def show_existing_files(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        kb: list[list[InlineKeyboardButton]] = [[InlineKeyboardButton(Path(p).name, callback_data=f"file_{i}")]
                                                for i, p in enumerate(self.existing_files)]
        kb.append([InlineKeyboardButton("↩️ Back", callback_data="back_to_source")])
        markup = InlineKeyboardMarkup(kb)
        await update.callback_query.edit_message_text("Select an existing file to process:", reply_markup=markup)
        return SELECT_EXISTING_FILE

    async def select_existing_file(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        q = update.callback_query; await q.answer()
        if q.data == "back_to_source":
            return await self.show_source_selection(update, context)
        idx = int(q.data.removeprefix("file_"))
        try:
            file_path = self.existing_files[idx]
        except IndexError:
            await q.edit_message_text("⚠️ File not found – pick again.")
            return await self.show_existing_files(update, context)
        await q.edit_message_text(f"⏳ Processing {Path(file_path).name} …")
        return await self._process_input_file(file_path, update, context)

    # ── new upload ───────────────────────────────────────────────
    async def file_received(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        tg_file = await update.message.document.get_file()
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=Path(update.message.document.file_name).suffix)
        await tg_file.download_to_file(tmp)
        tmp.close()
        return await self._process_input_file(tmp.name, update, context)

    # ── common loader ────────────────────────────────────────────
    async def _process_input_file(self, file_path: str, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        try:
            if file_path.lower().endswith(".xlsx"):
                self.df = pd.read_excel(file_path)
                self.df = _recover_link_column(self.df, file_path)
                self.df = _recover_photo_column(self.df)
            elif file_path.lower().endswith(".json"):
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                self.df = pd.json_normalize(data if isinstance(data, list) else next(v for v in data.values() if isinstance(v, list)))
            else:
                raise ValueError("Unsupported file type")
            self.file_path = file_path
        except Exception as exc:
            logger.exception("Load error")
            await context.bot.send_message(update.effective_chat.id, f"❌ Cannot read file: {exc}")
            return await self.show_source_selection(update, context)
        return await self._continue_after_loading(update, context)

    async def _continue_after_loading(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        self.all_columns = list(self.df.columns)
        # default visible columns
        self.selected_columns = [c for c in self.all_columns if c not in DEFAULT_EXCLUDED_COLUMNS]
        # years
        if "Year" in self.df.columns:
            self.all_years = sorted(self.df["Year"].dropna().unique().astype(int))
            self.selected_years = self.all_years.copy()
        # countries / collections
        if "Country" in self.df.columns:
            self.all_countries = sorted(self.df["Country"].dropna().unique())
            self.selected_countries = self.all_countries.copy()
        if "Collection" in self.df.columns:
            self.all_collections = sorted(self.df["Collection"].dropna().unique())
            self.selected_collections = self.all_collections.copy()
        return await self.show_column_selection(update, context)

    # ── column selection UI ──────────────────────────────────────
    async def show_column_selection(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        kb = [[InlineKeyboardButton(("✅ " if c in self.selected_columns else "❌ ") + c, callback_data=f"col_{c}")]
              for c in self.all_columns]
        kb.append([InlineKeyboardButton("✨ Confirm", callback_data="confirm_columns")])
        markup = InlineKeyboardMarkup(kb)
        if update.callback_query:
            await update.callback_query.edit_message_text("Select columns to include:", reply_markup=markup)
        else:
            await update.message.reply_text("Select columns to include:", reply_markup=markup)
        return SELECTING_COLUMNS

    async def toggle_column(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        q = update.callback_query; await q.answer()
        col = q.data.removeprefix("col_")
        if col in self.selected_columns:
            self.selected_columns.remove(col)
        else:
            self.selected_columns.append(col)
        return await self.show_column_selection(update, context)

    async def confirm_columns(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        q = update.callback_query; await q.answer()
        if not self.selected_columns:
            await q.edit_message_text("⚠️ Select at least one column.")
            return await self.show_column_selection(update, context)
        await q.edit_message_text("Column selection saved. ✓")
        # continue with year filtering if available
        if self.all_years:
            return await self.show_year_selection(update, context)
        if self.all_countries:
            return await self.show_country_selection(update, context)
        if self.all_collections:
            return await self.show_collection_selection(update, context)
        return await self.show_format_selection(update, context)

    # ── year filtering UI ───────────────────────────────────────
    async def show_year_selection(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        kb = [[InlineKeyboardButton(("✅ " if y in self.selected_years else "❌ ") + str(y), callback_data=f"year_{y}")]
              for y in self.all_years]
        kb.append([
            InlineKeyboardButton("✅ All", callback_data="year_all"),
            InlineKeyboardButton("❌ None", callback_data="year_none"),
        ])
        kb.append([InlineKeyboardButton("✨ Confirm", callback_data="confirm_years")])
        markup = InlineKeyboardMarkup(kb)
        await update.callback_query.edit_message_text("Filter by years:", reply_markup=markup)
        return FILTER_YEARS

    async def toggle_year(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        q = update.callback_query; await q.answer()
        data = q.data
        if data == "year_all":
            self.selected_years = self.all_years.copy()
        elif data == "year_none":
            self.selected_years = []
        else:
            yr = int(data.removeprefix("year_"))
            self.selected_years = [y for y in self.selected_years if y != yr] if yr in self.selected_years else self.selected_years + [yr]
        return await self.show_year_selection(update, context)

    async def confirm_years(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        q = update.callback_query; await q.answer()
        if not self.selected_years:
            await q.edit_message_text("⚠️ Select at least one year.")
            return await self.show_year_selection(update, context)
        await q.edit_message_text("Year filter saved. ✓")
        if self.all_countries:
            return await self.show_country_selection(update, context)
        if self.all_collections:
            return await self.show_collection_selection(update, context)
        return await self.show_format_selection(update, context)

    # ── country filtering UI ────────────────────────────────────
    async def show_country_selection(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        kb = [[InlineKeyboardButton(("✅ " if c in self.selected_countries else "❌ ") + c, callback_data=f"ctry_{i}")]
              for i, c in enumerate(self.all_countries)]
        kb.append([
            InlineKeyboardButton("✅ All", callback_data="ctry_all"),
            InlineKeyboardButton("❌ None", callback_data="ctry_none"),
        ])
        kb.append([InlineKeyboardButton("✨ Confirm", callback_data="confirm_countries")])
        markup = InlineKeyboardMarkup(kb)
        await update.callback_query.edit_message_text("Filter by countries:", reply_markup=markup)
        return FILTER_COUNTRIES

    async def toggle_country(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        q = update.callback_query; await q.answer()
        data = q.data
        if data == "ctry_all":
            self.selected_countries = self.all_countries.copy()
        elif data == "ctry_none":
            self.selected_countries = []
        else:
            idx = int(data.removeprefix("ctry_"))
            c = self.all_countries[idx]
            self.selected_countries = [x for x in self.selected_countries if x != c] if c in self.selected_countries else self.selected_countries + [c]
        return await self.show_country_selection(update, context)

    async def confirm_countries(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        q = update.callback_query; await q.answer()
        if not self.selected_countries:
            await q.edit_message_text("⚠️ Select at least one country.")
            return await self.show_country_selection(update, context)
        await q.edit_message_text("Country filter saved. ✓")
        if self.all_collections:
            return await self.show_collection_selection(update, context)
        return await self.show_format_selection(update, context)

    # ── collection filtering UI ─────────────────────────────────
    async def show_collection_selection(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        kb = [[InlineKeyboardButton(("✅ " if col in self.selected_collections else "❌ ") + (col[:23] + "…" if len(col) > 25 else col), callback_data=f"collec_{i}")]
              for i, col in enumerate(self.all_collections)]
        kb.append([
            InlineKeyboardButton("✅ All", callback_data="collec_all"),
            InlineKeyboardButton("❌ None", callback_data="collec_none"),
        ])
        kb.append([InlineKeyboardButton("✨ Confirm", callback_data="confirm_collections")])
        markup = InlineKeyboardMarkup(kb)
        await update.callback_query.edit_message_text("Filter by collections:", reply_markup=markup)
        return FILTER_COLLECTIONS

    async def toggle_collection(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        q = update.callback_query; await q.answer()
        data = q.data
        if data == "collec_all":
            self.selected_collections = self.all_collections.copy()
        elif data == "collec_none":
            self.selected_collections = []
        else:
            idx = int(data.removeprefix("collec_"))
            col = self.all_collections[idx]
            self.selected_collections = [x for x in self.selected_collections if x != col] if col in self.selected_collections else self.selected_collections + [col]
        return await self.show_collection_selection(update, context)

    async def confirm_collections(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        q = update.callback_query; await q.answer()
        if not self.selected_collections:
            await q.edit_message_text("⚠️ Select at least one collection.")
            return await self.show_collection_selection(update, context)
        await q.edit_message_text("Collection filter saved. ✓")
        return await self.show_format_selection(update, context)

    # ── format pick ─────────────────────────────────────────────
    async def show_format_selection(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        markup = InlineKeyboardMarkup([
            [InlineKeyboardButton("📊 Excel (XLSX)", callback_data="fmt_xlsx")],
            [InlineKeyboardButton("📄 PDF", callback_data="fmt_pdf")],
        ])
        if update.callback_query:
            await update.callback_query.edit_message_text("Choose output format:", reply_markup=markup)
        else:
            await update.message.reply_text("Choose output format:", reply_markup=markup)
        return SELECT_FORMAT

    async def select_format(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        q = update.callback_query; await q.answer()
        self.output_format = "pdf" if q.data == "fmt_pdf" else "xlsx"
        await q.edit_message_text("⏳ Processing your request …")
        await self.process_data(update, context)
        return ConversationHandler.END

    # ── core export ─────────────────────────────────────────────
    async def process_data(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
        chat_id = update.effective_chat.id
        note = await context.bot.send_message(chat_id, "⏳ crunching…")

        df = self.df.copy()
        if self.selected_years and "Year" in df.columns:
            df = df[df["Year"].isin(self.selected_years)]
        if self.selected_countries and "Country" in df.columns:
            df = df[df["Country"].isin(self.selected_countries)]
        if self.selected_collections and "Collection" in df.columns:
            df = df[df["Collection"].isin(self.selected_collections)]

        hide_photo = hide_link = False
        internal_cols = self.selected_columns.copy()
        for special in ("Photo", "Link"):
            if special not in internal_cols and special in df.columns:
                internal_cols.append(special)
                if special == "Photo": hide_photo = True
                if special == "Link": hide_link = True
        df = df[internal_cols]
        for c in ("Photo", "Link"):
            if c in df.columns:
                df[c] = df[c].astype(str).fillna("")

        chunks = split_dataframe(df, max_rows=500)
        for part, chunk in enumerate(chunks, 1):
            # Always create an Excel file first (even for PDF output)
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp.close()
            xlsx_path = tmp.name

            with pd.ExcelWriter(xlsx_path, engine="xlsxwriter") as xw:
                chunk.to_excel(xw, index=False, sheet_name="Sheet1")
                wb = xw.book
                ws = xw.sheets["Sheet1"]
                hdr_fmt = wb.add_format({"bold": True, "bg_color": "#DDEBF7", "align": "center", "valign": "vcenter"})
                ws.set_row(0, 25, hdr_fmt)

                # Set columns widths
                for i, name in enumerate(chunk.columns):
                    if name == "Photo":
                        # Make Photo column much wider - at least 25 characters
                        # This value matches what's used in stockgenerate.py (thumb_px / 7)
                        ws.set_column(i, i, 19)
                    else:
                        ws.set_column(i, i, max(15, len(name) + 2))

                # Set proper row height for all data rows (matching stockgenerate.py)
                row_h_pt = 170 * 0.75  # Same height calculation as in stockgenerate.py
                for r in range(1, len(chunk) + 1):
                    ws.set_row(r, row_h_pt)

                helper_link_col = None
                photo_col = None
                link_col = None

                # Handle Link column
                if "Link" in chunk.columns:
                    link_col = chunk.columns.get_loc("Link")
                    helper_link_col = link_col + 1
                    ws.write(0, helper_link_col, "Link_URL")
                    ws.set_column(helper_link_col, helper_link_col, None, None, {"hidden": True})

                    for r, url in enumerate(chunk["Link"], start=1):
                        if url and url.startswith("http"):
                            ws.write(r, helper_link_col, url)
                            ws.write_url(r, link_col, url, string="ALL PHOTOS LINK")
                        else:
                            ws.write(r, link_col, "")

                    if hide_link:
                        ws.set_column(link_col, link_col, None, None, {"hidden": True})

                # Handle Photo column - no text, only images
                if "Photo" in chunk.columns:
                    photo_col = chunk.columns.get_loc("Photo")

                    total_pics = sum(1 for u in chunk["Photo"] if u and u.startswith("http"))
                    done = 0

                    for r, url in enumerate(chunk["Photo"], start=1):
                        # Clear the cell (no text in Photo column)
                        ws.write(r, photo_col, "")

                        if url and url.startswith("http"):
                            # Use the improved insert_image_compat with larger scale
                            insert_image_compat(ws, r, photo_col, cached_png(url), x_scale=0.85, y_scale=0.85)
                            done += 1
                            if done % 10 == 0:
                                try:
                                    await note.edit_text(f"⏳ thumbnails {done}/{total_pics}")
                                except Exception:
                                    pass

                    if hide_photo:
                        ws.set_column(photo_col, photo_col, None, None, {"hidden": True})

                # If PDF output, set landscape orientation and fit width
                if self.output_format == "pdf":
                    # Set worksheet to landscape mode
                    ws.set_landscape()
                    # Fit the print to 1 page wide
                    ws.fit_to_pages(1, 0)  # 1 page wide, as many pages tall as needed
                    # Set paper size to A3 for better fit
                    ws.set_paper(8)  # 8 corresponds to A3 paper size
                    # Center horizontally
                    ws.center_horizontally()

            # Handle file size limits
            size = os.path.getsize(xlsx_path)
            if size > MAX_FILE_SIZE:
                # Try with higher compression first
                if photo_col is not None:
                    with pd.ExcelWriter(xlsx_path, engine="xlsxwriter") as xw:
                        chunk.to_excel(xw, index=False, sheet_name="Sheet1")
                        wb = xw.book
                        ws = xw.sheets["Sheet1"]

                        # Re-apply all the formatting but with higher compression
                        hdr_fmt = wb.add_format(
                            {"bold": True, "bg_color": "#DDEBF7", "align": "center", "valign": "vcenter"})
                        ws.set_row(0, 25, hdr_fmt)

                        # Set columns widths again
                        for i, name in enumerate(chunk.columns):
                            if name == "Photo":
                                ws.set_column(i, i, 19)
                            else:
                                ws.set_column(i, i, max(15, len(name) + 2))

                        # Set row heights again
                        for r in range(1, len(chunk) + 1):
                            ws.set_row(r, row_h_pt)

                        # Re-process the Link column
                        if "Link" in chunk.columns:
                            link_col = chunk.columns.get_loc("Link")
                            for r, url in enumerate(chunk["Link"], start=1):
                                if url and url.startswith("http"):
                                    ws.write_url(r, link_col, url, string="ALL PHOTOS LINK")
                                else:
                                    ws.write(r, link_col, "")

                            if hide_link:
                                ws.set_column(link_col, link_col, None, None, {"hidden": True})

                        # Re-process Photo column with higher compression
                        if photo_col is not None:
                            for r, url in enumerate(chunk["Photo"], start=1):
                                # Clear the cell (no text in Photo column)
                                ws.write(r, photo_col, "")

                                if url and url.startswith("http"):
                                    bio = cached_png(url, compression_level=9)  # Higher compression
                                    insert_image_compat(ws, r, photo_col, bio, x_scale=0.7,
                                                        y_scale=0.7)  # Smaller scale

                            if hide_photo:
                                ws.set_column(photo_col, photo_col, None, None, {"hidden": True})

                        # If PDF output, reapply the PDF settings
                        if self.output_format == "pdf":
                            ws.set_landscape()
                            ws.fit_to_pages(1, 0)
                            ws.set_paper(8)
                            ws.center_horizontally()

                # If still too large, then remove images
                if os.path.getsize(xlsx_path) > MAX_FILE_SIZE:
                    wb = openpyxl.load_workbook(xlsx_path)
                    ws = wb.active
                    for img in list(ws._images):
                        ws._images.remove(img)
                    wb.save(xlsx_path)
                    wb.close()

            # Different file handling based on format
            if self.output_format == "pdf":
                await note.edit_text("⏳ Converting to PDF...")

                # Create PDF filename
                pdf_path = xlsx_path.replace(".xlsx", ".pdf")

                try:
                    # Try using libreoffice for conversion (the most reliable method)
                    cmd = [
                        "libreoffice", "--headless", "--convert-to", "pdf",
                        "--outdir", os.path.dirname(xlsx_path), xlsx_path
                    ]
                    subprocess.run(cmd, check=True, timeout=60)

                    # If libreoffice conversion succeeded, send the PDF
                    if os.path.exists(pdf_path):
                        with open(pdf_path, "rb") as f:
                            payload = BytesIO(f.read())
                            payload.seek(0)

                        fname = f"filtered_stock_data_part{part}of{len(chunks)}.pdf" if len(
                            chunks) > 1 else "filtered_stock_data.pdf"
                        await context.bot.send_document(chat_id, payload, filename=fname)

                        # Clean up files
                        os.unlink(pdf_path)
                    else:
                        # Fallback if PDF wasn't created but no error was raised
                        await note.edit_text("⚠️ PDF conversion failed, sending Excel file instead")
                        with open(xlsx_path, "rb") as f:
                            payload = BytesIO(f.read())
                            payload.seek(0)

                        fname = f"filtered_stock_data_part{part}of{len(chunks)}.xlsx" if len(
                            chunks) > 1 else "filtered_stock_data.xlsx"
                        await context.bot.send_document(chat_id, payload, filename=fname)

                except Exception as e:
                    # If PDF conversion fails, let the user know and send Excel instead
                    logger.error(f"PDF conversion error: {e}")
                    await note.edit_text(f"⚠️ PDF conversion failed ({str(e)[:30]}...), sending Excel file instead")

                    with open(xlsx_path, "rb") as f:
                        payload = BytesIO(f.read())
                        payload.seek(0)

                    fname = f"filtered_stock_data_part{part}of{len(chunks)}.xlsx" if len(
                        chunks) > 1 else "filtered_stock_data.xlsx"
                    await context.bot.send_document(chat_id, payload, filename=fname)

            else:  # XLSX format (original behavior)
                with open(xlsx_path, "rb") as f:
                    payload = BytesIO(f.read())
                    payload.seek(0)

                fname = f"filtered_stock_data_part{part}of{len(chunks)}.xlsx" if len(
                    chunks) > 1 else "filtered_stock_data.xlsx"
                await context.bot.send_document(chat_id, payload, filename=fname)

            # Clean up the Excel file
            os.unlink(xlsx_path)

        await note.edit_text("✅ done")

    async def cancel(self, update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
        await update.message.reply_text("Operation cancelled. Use /start to begin again.", reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END


    # ────────────────────────────────────────────────────────────────
    # main()
    # ────────────────────────────────────────────────────────────────

def main() -> None:
    application = Application.builder().token("7315690900:AAE8r-wipNaa8LXN-TqX8PU6-xsRI_BuqR8").build()

    bot = StockSelectorBot()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", bot.start)],
        states={
            SELECT_SOURCE: [CallbackQueryHandler(bot.select_source, pattern=r"^source_")],
            SELECT_FILE: [MessageHandler(filters.Document.ALL, bot.file_received)],
            SELECT_EXISTING_FILE: [CallbackQueryHandler(bot.select_existing_file, pattern=r"^(file_|back_)")],
            SELECTING_COLUMNS: [
                CallbackQueryHandler(bot.toggle_column, pattern=r"^col_"),
                CallbackQueryHandler(bot.confirm_columns, pattern=r"^confirm_columns$")
            ],
            FILTER_YEARS: [
                CallbackQueryHandler(bot.toggle_year, pattern=r"^(year_|year_all|year_none)"),
                CallbackQueryHandler(bot.confirm_years, pattern=r"^confirm_years$")
            ],
            FILTER_COUNTRIES: [
                CallbackQueryHandler(bot.toggle_country, pattern=r"^(ctry_|ctry_all|ctry_none)"),
                CallbackQueryHandler(bot.confirm_countries, pattern=r"^confirm_countries$")
            ],
            FILTER_COLLECTIONS: [
                CallbackQueryHandler(bot.toggle_collection, pattern=r"^(collec_|collec_all|collec_none)"),
                CallbackQueryHandler(bot.confirm_collections, pattern=r"^confirm_collections$")
            ],
            SELECT_FORMAT: [CallbackQueryHandler(bot.select_format, pattern=r"^fmt_")],
        },
        fallbacks=[CommandHandler("cancel", bot.cancel)],
    )

    application.add_handler(conv)
    application.run_polling()


if __name__ == "__main__":
    main()